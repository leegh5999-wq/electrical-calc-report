"""Generic sheet → JSON extractor for 전기계산서 lookup tables.

Set SRC (env var ECR_SRC or edit below) to the workbook path before running.

Two modes:
  --preview --all          # peek first 8 rows of each Phase-0 table (no write)
  --preview <sheet>        # peek a single sheet
  --all                    # extract Phase-0 6 tables to data/tables/
  <sheet> [--header-row N] # extract one sheet, optionally overriding header

Output JSON per table:
  source, sheet, dimensions, header_row, headers, rows (list of dicts), raw (2D)
"""
from __future__ import annotations

import argparse
import json
import os
import sys
from pathlib import Path

import openpyxl

# 원본 엑셀 경로. 환경변수 ECR_SRC 로 덮어쓸 수 있음.
SRC = Path(os.environ.get("ECR_SRC", r"C:\path\to\전기계산서.xlsx"))
PROJ = Path(__file__).resolve().parent.parent
OUT_DIR = PROJ / "data" / "tables"

# (sheet, header_row, output_slug). Headers chosen from preview.
# Where the sheet has no clean header row (multi-table layout, prose docs, or
# header-less data), header_row points at the closest usable line and the JS
# side should consume `raw` 2D directly using column positions.
COMMON_TABLES = [
    ("설계조건",      1, "design_conditions"),   # prose-form doc; raw 2D is primary
    ("DF",          6, "derating_factor"),       # KECG-17 cable derating (NOT 수용률); r6 = numeric labels
    ("조명기구",      3, "lighting_fixtures"),    # no real header; consume by column position
    ("CB",          6, "cb"),                    # first sub-table header at r6; multiple sub-tables
    ("CABLE DATA",  2, "cable_data"),            # 4-5 sub-tables side-by-side; r2 has 공칭단면적/외경/단면적
    ("장비일람",      3, "equipment_list"),       # clean header at r3; section rows interspersed
    ("수용률",        4, "demand_factor"),        # building-type × load-type demand factor [%]
]


def cell_value(v):
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v


def read_raw(ws):
    max_row = ws.max_row or 0
    max_col = ws.max_column or 0
    raw = []
    for r in range(1, max_row + 1):
        row = [cell_value(ws.cell(row=r, column=c).value) for c in range(1, max_col + 1)]
        raw.append(row)
    while raw and all(c is None for c in raw[-1]):
        raw.pop()
    return raw, max_row, max_col


def preview(ws, sheet_name: str, n: int = 8) -> None:
    raw, mr, mc = read_raw(ws)
    print(f"\n── {sheet_name}  ({mr}r × {mc}c, trimmed → {len(raw)}r) ─────────")
    for i, row in enumerate(raw[:n], 1):
        cells = []
        for j, v in enumerate(row[:12], 1):  # cap at 12 cols for readability
            if v is None:
                cells.append("·")
            else:
                s = str(v)
                cells.append(s if len(s) <= 14 else s[:13] + "…")
        print(f"  r{i:>2}: {' | '.join(cells)}")
    if mc > 12:
        print(f"       (+ {mc - 12} more cols)")


def extract(ws, sheet_name: str, header_row: int):
    raw, mr, mc = read_raw(ws)
    headers_raw = raw[header_row - 1] if 1 <= header_row <= len(raw) else []
    headers, seen = [], {}
    for i, h in enumerate(headers_raw):
        name = (str(h).strip() if h is not None and str(h).strip() else f"col_{i + 1}")
        k = seen.get(name, 0)
        unique_name = name if k == 0 else f"{name}_{k + 1}"
        seen[name] = k + 1
        headers.append(unique_name)

    rows = []
    for r in raw[header_row:]:
        if all(c is None for c in r):
            continue
        rows.append({headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))})

    return {
        "source": SRC.name,
        "sheet": sheet_name,
        "dimensions": {"max_row": mr, "max_col": mc, "data_rows": len(rows)},
        "header_row": header_row,
        "headers": headers,
        "rows": rows,
        "raw": raw,
    }


def main() -> int:
    ap = argparse.ArgumentParser()
    ap.add_argument("sheet", nargs="?")
    ap.add_argument("--header-row", type=int, default=1)
    ap.add_argument("--out", type=Path, default=None)
    ap.add_argument("--all", action="store_true")
    ap.add_argument("--preview", action="store_true")
    ap.add_argument("--src", type=Path, default=SRC)
    args = ap.parse_args()

    print(f"Loading: {args.src.name}")
    wb = openpyxl.load_workbook(args.src, data_only=True, read_only=False)

    if args.preview:
        targets = [t[0] for t in COMMON_TABLES] if args.all else ([args.sheet] if args.sheet else [])
        if not targets:
            ap.error("--preview needs a sheet or --all")
        for sn in targets:
            if sn not in wb.sheetnames:
                print(f"  MISSING: {sn}", file=sys.stderr)
                continue
            preview(wb[sn], sn)
        return 0

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    if args.all:
        tasks = COMMON_TABLES
    elif args.sheet:
        slug = args.sheet.replace(" ", "_").replace("/", "_")
        tasks = [(args.sheet, args.header_row, slug)]
    else:
        ap.error("provide sheet, --all, or --preview")

    for sheet_name, header_row, slug in tasks:
        if sheet_name not in wb.sheetnames:
            print(f"  SKIP {sheet_name}: not found", file=sys.stderr)
            continue
        data = extract(wb[sheet_name], sheet_name, header_row)
        out_path = args.out if args.out else OUT_DIR / f"{slug}.json"
        out_path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding="utf-8")
        size_kb = out_path.stat().st_size / 1024
        print(f"  {sheet_name:14s} → {out_path.name:28s} "
              f"[{data['dimensions']['max_row']}r×{data['dimensions']['max_col']}c, "
              f"data:{data['dimensions']['data_rows']}, {size_kb:.0f} KB]")
    return 0


if __name__ == "__main__":
    sys.exit(main())
