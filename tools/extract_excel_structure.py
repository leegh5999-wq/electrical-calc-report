"""Extract sheet structure + cross-sheet reference graph from an electrical
calc workbook, so we can plan the web-app calculator order and JSON schemas.

Set SRC (env var ECR_SRC or edit below) to the workbook path before running.

Outputs:
  - docs/excel_structure.json  : full machine-readable dump
  - stdout                     : concise human summary
"""
from __future__ import annotations

import json
import os
import re
import sys
from collections import Counter, defaultdict
from pathlib import Path

import openpyxl

# 원본 엑셀 경로. 환경변수 ECR_SRC 로 덮어쓸 수 있음.
SRC = Path(os.environ.get("ECR_SRC", r"C:\path\to\전기계산서.xlsx"))
OUT_DIR = Path(__file__).resolve().parent.parent / "docs"
OUT_JSON = OUT_DIR / "excel_structure.json"

# Match references like  '시트 이름'!$A$1  or  시트명!B2  (with optional quotes,
# optional absolute markers). Sheet name may contain spaces/Korean/() etc.
REF_RE = re.compile(r"(?:'([^']+)'|([A-Za-z가-힣0-9_\-\. \(\)]+))!\$?[A-Z]+\$?\d+")


def sheet_refs_in_formula(formula: str) -> list[str]:
    if not isinstance(formula, str) or "!" not in formula:
        return []
    hits = []
    for m in REF_RE.finditer(formula):
        name = m.group(1) or m.group(2)
        if name:
            hits.append(name.strip())
    return hits


def main() -> int:
    if not SRC.exists():
        print(f"ERROR: source not found: {SRC}", file=sys.stderr)
        return 2

    print(f"Loading: {SRC.name}  ({SRC.stat().st_size:,} bytes)")
    wb = openpyxl.load_workbook(SRC, data_only=False, read_only=False)

    sheet_names = wb.sheetnames
    print(f"Sheets: {len(sheet_names)}")

    # Defined names (workbook-level + sheet-level)
    defined_names: list[dict] = []
    try:
        for dn in wb.defined_names.definedName if hasattr(wb.defined_names, "definedName") else []:
            defined_names.append({"name": dn.name, "value": dn.value, "scope": getattr(dn, "localSheetId", None)})
    except Exception:
        # openpyxl 3.1: defined_names behaves like a dict
        try:
            for nm in wb.defined_names:
                d = wb.defined_names[nm]
                defined_names.append({"name": nm, "value": getattr(d, "value", str(d))})
        except Exception as e:
            print(f"  (defined_names enumeration failed: {e})")

    sheets_info: list[dict] = []
    edges: list[tuple[str, str, int]] = []  # (from_sheet, to_sheet, count)
    cross_edge_counter: dict[tuple[str, str], int] = defaultdict(int)

    for sn in sheet_names:
        ws = wb[sn]
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0
        merged_count = len(ws.merged_cells.ranges)

        cell_count = 0
        formula_count = 0
        value_count = 0
        external_refs = 0
        ref_targets_local: Counter[str] = Counter()
        sample_formulas: list[tuple[str, str]] = []

        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if v is None:
                    continue
                cell_count += 1
                if isinstance(v, str) and v.startswith("="):
                    formula_count += 1
                    for target in sheet_refs_in_formula(v):
                        # skip self-references
                        if target == sn:
                            continue
                        ref_targets_local[target] += 1
                        cross_edge_counter[(sn, target)] += 1
                    if "[" in v and "]" in v:
                        external_refs += 1
                    if len(sample_formulas) < 5 and "!" in v:
                        sample_formulas.append((cell.coordinate, v[:160]))
                else:
                    value_count += 1

        sheets_info.append({
            "name": sn,
            "max_row": max_row,
            "max_col": max_col,
            "merged_ranges": merged_count,
            "cells_non_empty": cell_count,
            "formulas": formula_count,
            "values": value_count,
            "external_ref_formulas": external_refs,
            "refs_to_other_sheets": dict(ref_targets_local),
            "sample_cross_sheet_formulas": sample_formulas,
        })

    for (a, b), c in cross_edge_counter.items():
        edges.append((a, b, c))

    # Build dependency in-degree to suggest topological order
    in_deg: Counter[str] = Counter({sn: 0 for sn in sheet_names})
    out_deg: Counter[str] = Counter({sn: 0 for sn in sheet_names})
    for a, b, _ in edges:
        if b in in_deg:
            in_deg[b] += 1
        if a in out_deg:
            out_deg[a] += 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    out = {
        "source": str(SRC),
        "sheet_order": sheet_names,
        "defined_names": defined_names,
        "sheets": sheets_info,
        "cross_sheet_edges": [
            {"from": a, "to": b, "count": c} for a, b, c in sorted(edges, key=lambda x: -x[2])
        ],
        "in_degree": dict(in_deg),
        "out_degree": dict(out_deg),
    }
    OUT_JSON.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote: {OUT_JSON}")

    # --- human summary ---
    print("\n=== Sheet list (in workbook order) ===")
    for i, sn in enumerate(sheet_names, 1):
        info = next(s for s in sheets_info if s["name"] == sn)
        print(f"  {i:2d}. {sn}  "
              f"[{info['max_row']}r x {info['max_col']}c, "
              f"형식={info['formulas']}, 값={info['values']}, "
              f"외부참조={info['external_ref_formulas']}, "
              f"in={in_deg[sn]}, out={out_deg[sn]}]")

    print("\n=== Top cross-sheet edges (from → to : count) ===")
    for a, b, c in sorted(edges, key=lambda x: -x[2])[:25]:
        print(f"  {a!r:35s} → {b!r:35s} : {c}")

    if defined_names:
        print(f"\n=== Defined names ({len(defined_names)}) ===")
        for d in defined_names[:20]:
            print(f"  {d.get('name')}  =  {d.get('value')}")

    # Suggested topological build order (sources first)
    print("\n=== Suggested build order (low in-degree first) ===")
    ordered = sorted(sheet_names, key=lambda s: (in_deg[s], -out_deg[s], sheet_names.index(s)))
    for i, sn in enumerate(ordered, 1):
        print(f"  {i:2d}. {sn}  (in={in_deg[sn]}, out={out_deg[sn]})")

    return 0


if __name__ == "__main__":
    sys.exit(main())
